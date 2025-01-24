import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook, load_workbook
import math
import os


class DegerlendirmeIslemleri:
    def __init__(self):
        # DataFrame'ler
        self.df_ogrnotlar = None   # Öğrenci Notları
        self.df_deg = None         # Ders Çıktısı
        self.df_prgtablo = None    # Program Çıktıları

        # Hesaplama parametreleri
        self.ders_sayisi = 0
        self.ders_cikti_yuzdeleri = []  # örn. [0.34, 0.26, 0.40] vs.

        # İçsel tablolar (hesaplanan)
        self.tablo3df = None
        self.tablo4_dfler = {}
        self.tablo5_dfler = {}

    def truncate(self, number, digits):
        stepper = 10.0 ** digits
        return math.trunc(stepper * number) / stepper

   
    def _convert_numeric_columns(self, df):
        if df.shape[1] > 1:
            # İlk sütun = metin, geri kalan sütunlar numeric
            df.iloc[:, 1:] = df.iloc[:, 1:].apply(
                pd.to_numeric, errors="coerce"
            )
            # Kontrol: NaN var mı?
            if df.iloc[:, 1:].isna().any().any():
                messagebox.showwarning(
                    "Uyarı",
                    "Bazı hücreler sayısal olmayan değerler içeriyor. "
                    "Bu değerler NaN (boş) olarak işlenecek!"
                )
        return df

  
    def _check_range(self, df, min_val, max_val, tablo_adi=""):
        if df.shape[1] <= 1:
            return
        numeric_part = df.iloc[:, 1:]  # ilk sütun haricindeki sütunlar
        # True/False tablosu
        out_of_range_mask = (numeric_part < min_val) | (numeric_part > max_val)
        if out_of_range_mask.any().any():
            raise ValueError(
                f"{tablo_adi} tablosunda {min_val}-{max_val} aralığı dışında değer(ler) var!"
            )

  
    def sec_ogr_notlar_dosyasi(self, file_path):
        """
        Öğrenci Notları dosyası -> 0-100 aralığı
        """
        df_temp = pd.read_excel(file_path)
        df_temp.columns = df_temp.columns.str.strip()
        df_temp = self._convert_numeric_columns(df_temp)

        # 0-100 aralığı kontrolü
        self._check_range(df_temp, 0, 100, tablo_adi="Öğrenci Notları")

        self.df_ogrnotlar = df_temp
        col_count = df_temp.shape[1]
        if col_count <= 1:
            raise ValueError(
                "Öğrenci Notları tablosunda en az 1 not sütunu olmalı!"
            )
        self.ders_sayisi = col_count - 1

    def sec_degerlendirmeler_dosyasi(self, file_path):
        df_temp = pd.read_excel(file_path, header=1)
        wb = load_workbook(file_path)
        sheet = wb.active

        # sheet[1] = ilk satır (Python'da index 1, Excel'de satır 1)
        self.ders_cikti_yuzdeleri = list(map(lambda x: float(
            x)/100, [cell.value for cell in sheet[1]][1:]))  # ders cikti yuzdeleri okunuyor
        df_temp.columns = df_temp.columns.str.strip()

        df_temp = self._convert_numeric_columns(df_temp)

        # 0-1 aralığı kontrolü
        self._check_range(df_temp, 0, 1, tablo_adi="Ders Çıktısı")

        self.df_deg = df_temp

    def sec_prgcikti_dosyasi(self, file_path):
        """
        Program Çıktıları dosyası -> 0-1 aralığı
        """
        df_temp = pd.read_excel(file_path, header=1)
        df_temp = self._convert_numeric_columns(df_temp)

        # 0-1 aralığı kontrolü
        self._check_range(df_temp, 0, 1, tablo_adi="Program Çıktıları")

        self.df_prgtablo = df_temp

   
    def create_ogr_table(self, not_isimleri, not_oranlari):
        """
        Ör: not_isimleri=["not1","not2"], not_oranlari=[34,66]
        => df_ogrnotlar columns = ["Öğrenci","not1","not2"]
        => ders_sayisi=2, ders_cikti_yuzdeleri=[0.34,0.66]
        """
        columns = ["Öğrenci"] + not_isimleri
        self.df_ogrnotlar = pd.DataFrame(columns=columns)
        self.ders_sayisi = len(not_isimleri)
        self.ders_cikti_yuzdeleri = [val / 100.0 for val in not_oranlari]

    def create_deg_table_from_ogr(self):
        """
        Ders Çıktısı tablosu: ["Ders Çıktısı"] + ogrnotlar.columns[1:]
        """
        if self.df_ogrnotlar is None:
            raise ValueError("Önce Öğrenci Notları oluşturun veya yükleyin!")
        not_sutunlari = list(self.df_ogrnotlar.columns[1:])
        columns = ["Ders Çıktısı"] + not_sutunlari
        self.df_deg = pd.DataFrame(columns=columns)

    def create_prg_table_from_deg(self):
        """
        Program Çıktıları: ["Program Çıktısı"] + (df_deg'in satır sayısı kadar ilk sütun)
        """
        if self.df_deg is None or self.df_deg.shape[0] == 0:
            raise ValueError("Ders Çıktısı tablosu yok veya boş!")
        # df_deg'in ilk sütunundaki değerler
        ders_adlari = self.df_deg.iloc[:, 0].tolist()
        columns = ["Program Çıktısı"] + ders_adlari
        self.df_prgtablo = pd.DataFrame(columns=columns)

    def olustur_tablo3df(self):
        """
        Tablo3 => Sadece hafızada kullanılır. (Excel'e kaydedilmez)
        Ders Çıktısı tablosunun (df_deg) her satırındaki not değerlerini (float) -> ders_cikti_yuzdeleri ile çarpar.
        Son sütun "Toplam" eklenir.
        """
        if self.df_deg is None or not self.ders_cikti_yuzdeleri:
            self.tablo3df = None
            return

        # ["Ders Çıktısı","not1","not2",...]
        columns = list(self.df_deg.columns)
        self.tablo3df = pd.DataFrame(columns=columns + ["Toplam"])

        for _, deg_row in self.df_deg.iterrows():
            row_data = []
            # İlk sütun = Ders Çıktısı metin
            row_data.append(deg_row.iloc[0])
            toplam = 0
            for i in range(1, 1 + self.ders_sayisi):
                val = deg_row.iloc[i]
                ratio = self.ders_cikti_yuzdeleri[i - 1]
                if pd.isna(val):
                    deger = float("nan")
                else:
                    deger = val * ratio
                if pd.notna(deger):
                    toplam += deger
                row_data.append(deger)
            row_data.append(toplam)
            self.tablo3df.loc[len(self.tablo3df)] = row_data

    def olustur_tablo4_dfler(self):
        """
        Her öğrenci (df_ogrnotlar satırı) için tablo3'teki her satır (Ders Çıktısı) ile çarpıp
        "Toplam, MAX, % Başarı" hesaplar.
        """
        self.tablo4_dfler = {}
        if (self.df_ogrnotlar is None) or (self.tablo3df is None):
            return

        for _, ogr_row in self.df_ogrnotlar.iterrows():
            ogr_adi = str(ogr_row.iloc[0])  # Öğrenci ismi
            kolonlar = list(self.df_deg.columns) + \
                ["Toplam", "MAX", "% Başarı"]
            df_tablo4 = pd.DataFrame(columns=kolonlar)

            for _, ders_row in self.tablo3df.iterrows():
                row_data = []
                row_data.append(ders_row.iloc[0])  # Ders Çıktısı
                toplam = 0
                maxtoplam = 0

                # ders_sayisi adet not
                for i in range(1, 1 + self.ders_sayisi):
                    # öğrencinin notu (NaN olabilir)
                    val_ogr = ogr_row.iloc[i]
                    val_ders = ders_row.iloc[i]  # tablo3'teki (NaN olabilir)
                    if pd.isna(val_ogr) or pd.isna(val_ders):
                        deger = float("nan")
                        maxdeger = float("nan")
                    else:
                        deger = val_ogr * val_ders
                        maxdeger = val_ders * 100
                    row_data.append(deger)

                    if pd.notna(deger):
                        toplam += deger
                    if pd.notna(maxdeger):
                        maxtoplam += maxdeger

                row_data.append(toplam)
                row_data.append(maxtoplam)

                if maxtoplam != 0:
                    basari = self.truncate(100 * toplam / maxtoplam, 2)
                else:
                    basari = float("nan")

                row_data.append(basari)
                df_tablo4.loc[len(df_tablo4)] = row_data

            self.tablo4_dfler[ogr_adi] = df_tablo4

    def olustur_tablo5_dfler(self):
        """
        Her öğrenci için tablo4'teki "% Başarı" sütununu
        program çıktı tablosundaki (NaN olabilir) değerlerle çarpar.
        """
        self.tablo5_dfler = {}
        if not self.tablo4_dfler or self.df_prgtablo is None:
            return

        for ogr_adi, df_tablo4 in self.tablo4_dfler.items():
            basari_list = df_tablo4["% Başarı"].tolist()  # NaN olabilir
            kolonlar = ["Prg Çıktı"] + basari_list + ["Başarı Oranı"]
            df_tablo5 = pd.DataFrame(columns=kolonlar)

            for _, prg_row in self.df_prgtablo.iterrows():
                row_data = []
                toplam = 0
                maxtoplam = 0

                row_data.append(prg_row.iloc[0])  # Program Çıktısı
                for i in range(len(basari_list)):
                    val_bsr = basari_list[i]   # NaN olabilir
                    val_prg = prg_row.iloc[i + 1]  # NaN olabilir
                    if pd.isna(val_bsr) or pd.isna(val_prg):
                        deger = float("nan")
                    else:
                        deger = val_bsr * val_prg
                    row_data.append(deger)

                    if pd.notna(deger):
                        toplam += deger
                        if pd.notna(val_prg):
                            maxtoplam += val_prg * 100

                if maxtoplam != 0:
                    basari_orani = self.truncate(100 * toplam / maxtoplam, 2)
                else:
                    basari_orani = float("nan")

                row_data.append(basari_orani)
                df_tablo5.loc[len(df_tablo5)] = row_data

            self.tablo5_dfler[ogr_adi] = df_tablo5

    def tum_islemleri_yap(self):
        self.olustur_tablo3df()
        self.olustur_tablo4_dfler()
        self.olustur_tablo5_dfler()

    
    def yaz_tablo4(self, output_file="tablo4.xlsx"):
        if not self.tablo4_dfler:
            return
        wb = Workbook()
        sh = wb.active
        sh.title = "Tablo4"

        row_number = 1
        for ogr_adi, df_tablo4 in self.tablo4_dfler.items():
            sh.cell(row=row_number, column=1, value="Tablo4")
            sh.merge_cells(
                start_row=row_number, end_row=row_number, start_column=2, end_column=3
            )
            sh.cell(
                row=row_number, column=2, value=f"Öğrenci {ogr_adi} için"
            )
            row_number += 1

            for c_idx, col_name in enumerate(df_tablo4.columns, start=1):
                sh.cell(row=row_number, column=c_idx, value=col_name)

            for row_data in df_tablo4.itertuples(index=False):
                row_number += 1
                for c_idx, value in enumerate(row_data, start=1):
                    cell_val = "" if pd.isna(value) else value
                    sh.cell(row=row_number, column=c_idx, value=cell_val)
            row_number += 2

        wb.save(output_file)

    def yaz_tablo5(self, output_file="tablo5.xlsx"):
        if not self.tablo5_dfler:
            return
        wb = Workbook()
        sh = wb.active
        sh.title = "Tablo5"

        row_number = 1
        for ogr_adi, df_tablo5 in self.tablo5_dfler.items():
            sh.cell(row=row_number, column=1, value=f"Öğrenci {ogr_adi} için")
            row_number += 1

            sh.cell(row=row_number, column=1, value="Tablo5")
            if self.ders_sayisi:
                sh.merge_cells(
                    start_row=row_number,
                    end_row=row_number,
                    start_column=2,
                    end_column=2 + self.ders_sayisi,
                )
            sh.cell(row=row_number, column=2, value="Ders Çıktısı")
            row_number += 1

            for c_idx, col_name in enumerate(df_tablo5.columns, start=1):
                sh.cell(row=row_number, column=c_idx, value=col_name)

            for row_data in df_tablo5.itertuples(index=False):
                row_number += 1
                for c_idx, value in enumerate(row_data, start=1):
                    cell_val = "" if pd.isna(value) else value
                    sh.cell(row=row_number, column=c_idx, value=cell_val)
            row_number += 2

        wb.save(output_file)



class DegerlendirmeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Değerlendirme Uygulaması")
        self.geometry("700x600")

        self.islem = DegerlendirmeIslemleri()

        # Dosya yolları
        self.ogr_file_path = None
        self.deg_file_path = None
        self.prg_file_path = None

        # Etiketler
        self.lbl_ogr = tk.Label(
            self, text="Öğrenci Notları Dosyası: Henüz Seçilmedi"
        )
        self.lbl_deg = tk.Label(
            self, text="Değerlendirmeler Dosyası: Henüz Seçilmedi"
        )
        self.lbl_prg = tk.Label(
            self, text="Program Çıktıları Dosyası: Henüz Seçilmedi"
        )

        # Butonlar: Dosya Seç
        self.btn_ogr = tk.Button(
            self, text="Öğrenci Notları Seç", command=self.choose_ogr_not_file
        )
        self.btn_deg = tk.Button(
            self, text="Değerlendirmeler Seç", command=self.choose_degerlendirme_file
        )
        self.btn_prg = tk.Button(
            self, text="Program Çıktıları Seç", command=self.choose_prg_cikti_file
        )

        # Ders Adı metin kutusu (EKLENDİ)
        self.lbl_dersadi = tk.Label(self, text="Ders Adı:")
        self.txt_dersadi = tk.Entry(self)

        # Butonlar: Hesaplama & Kaydet
        self.btn_run = tk.Button(
            self, text="Hesaplamaları Yap", command=self.run_calculations
        )
        self.btn_save = tk.Button(
            self, text="Tabloları Kaydet (4-5)", command=self.save_tables
        )

        # Tabloları Görüntüleme
        self.btn_show_ogr = tk.Button(
            self, text="Öğrenci Notları Tablosu", command=self.show_ogrnotlar_table
        )
        self.btn_show_deg = tk.Button(
            self, text="Değerlendirmeler Tablosu", command=self.show_degerlendirmeler_table
        )
        self.btn_show_prg = tk.Button(
            self, text="Program Çıktıları Tablosu", command=self.show_prgciktisi_table
        )

        # Layout
        self.lbl_ogr.pack(pady=5)
        self.btn_ogr.pack(pady=5)

        self.lbl_deg.pack(pady=5)
        self.btn_deg.pack(pady=5)

        self.lbl_prg.pack(pady=5)
        self.btn_prg.pack(pady=5)

        # Ders Adı label & entry (EKLENDİ)
        self.lbl_dersadi.pack(pady=5)
        self.txt_dersadi.pack(pady=5)

        self.btn_run.pack(pady=5)
        self.btn_save.pack(pady=5)

        self.btn_show_ogr.pack(pady=5)
        self.btn_show_deg.pack(pady=5)
        self.btn_show_prg.pack(pady=5)

    
    def choose_ogr_not_file(self):
        file_path = filedialog.askopenfilename(
            title="Öğrenci Notları Dosyası Seç",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if file_path:
            try:
                self.islem.sec_ogr_notlar_dosyasi(file_path)
                self.ogr_file_path = file_path
                self.lbl_ogr.config(
                    text=f"Öğrenci Notları Dosyası: {file_path}"
                )
                messagebox.showinfo(
                    "Bilgi", "Öğrenci Notları dosyası yüklendi.")
            except Exception as e:
                messagebox.showerror("Hata", f"Yükleme hatası: {e}")

    def choose_degerlendirme_file(self):
        file_path = filedialog.askopenfilename(
            title="Ders Çıktısı Dosyası Seç",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if file_path:
            try:
                self.islem.sec_degerlendirmeler_dosyasi(file_path)
                self.deg_file_path = file_path
                self.lbl_deg.config(
                    text=f"Ders Çıktısı Dosyası: {file_path}"
                )
                messagebox.showinfo("Bilgi", "Ders Çıktısı dosyası yüklendi.")
            except Exception as e:
                messagebox.showerror("Hata", f"Yükleme hatası: {e}")

    def choose_prg_cikti_file(self):
        file_path = filedialog.askopenfilename(
            title="Program Çıktıları Dosyası Seç",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if file_path:
            try:
                self.islem.sec_prgcikti_dosyasi(file_path)
                self.prg_file_path = file_path
                self.lbl_prg.config(
                    text=f"Program Çıktıları Dosyası: {file_path}"
                )
                messagebox.showinfo(
                    "Bilgi", "Program Çıktıları dosyası yüklendi."
                )
            except Exception as e:
                messagebox.showerror("Hata", f"Yükleme hatası: {e}")

    
    def create_ogr_table_popup(self):
        popup1 = tk.Toplevel(self)
        popup1.title("Öğrenci Notları Oluştur - Adım 1/2")
        popup1.geometry("400x200")

        lbl_info = tk.Label(
            popup1, text="Not isimlerini virgülle giriniz (Ör: not1, not2, not3)"
        )
        lbl_info.pack(pady=10)

        txt_notlar = tk.Text(popup1, height=4, width=30)
        txt_notlar.pack(pady=5)

        def on_ok():
            val = txt_notlar.get("1.0", tk.END).strip()
            if not val:
                messagebox.showerror("Hata", "Not isimleri girmediniz.")
                return
            not_names = [x.strip() for x in val.split(",") if x.strip()]
            if len(not_names) == 0:
                messagebox.showerror("Hata", "Geçersiz not isimleri!")
                return

            popup1.destroy()
            # Şimdi 2. popup => oranlar
            self._ask_oranlar_popup(not_names)

        btn = tk.Button(popup1, text="Devam", command=on_ok)
        btn.pack(pady=10)

    def _ask_oranlar_popup(self, not_names):
        popup2 = tk.Toplevel(self)
        popup2.title("Öğrenci Notları Oluştur - Adım 2/2")
        popup2.geometry("400x200")

        lbl_info = tk.Label(
            popup2,
            text=(
                f"{len(not_names)} not ismi bulduk.\n"
                "Sırasıyla oranlarını virgülle giriniz (toplam 100).\n"
                "Örnek: 34, 45, 21"
            ),
        )
        lbl_info.pack(pady=10)

        txt_oran = tk.Text(popup2, height=4, width=30)
        txt_oran.pack(pady=5)

        def on_ok2():
            val_oran = txt_oran.get("1.0", tk.END).strip()
            if not val_oran:
                messagebox.showerror("Hata", "Oranlar boş!")
                return
            oran_list = [x.strip() for x in val_oran.split(",") if x.strip()]

            if len(oran_list) != len(not_names):
                messagebox.showerror(
                    "Hata",
                    f"{len(not_names)} not ismi var ama {len(oran_list)} oran girdiniz!",
                )
                return

            try:
                oranlar = [float(o) for o in oran_list]
            except ValueError:
                messagebox.showerror("Hata", "Oranlar sayısal olmalı!")
                return

            if abs(sum(oranlar) - 100.0) > 1e-5:
                messagebox.showerror("Hata", "Oranların toplamı 100 değil!")
                return

            try:
                self.islem.create_ogr_table(not_names, oranlar)
                self.lbl_ogr.config(text="Öğrenci Notları: (Yeni Oluşturuldu)")
                messagebox.showinfo(
                    "Bilgi", "Öğrenci Notları tablosu oluşturuldu."
                )
                popup2.destroy()
            except Exception as e:
                messagebox.showerror("Hata", str(e))

        btn2 = tk.Button(popup2, text="Oluştur", command=on_ok2)
        btn2.pack(pady=10)

    
    def show_ogrnotlar_table(self):
        if self.islem.df_ogrnotlar is None:
            self.create_ogr_table_popup()  # 2 aşamalı
            return
        self._show_table_window(
            df_attr_name="df_ogrnotlar",
            title="Öğrenci Notları Tablosu",
            add_button_text="Öğrenci Ekle",
            save_callback=self.save_new_student_row,
        )

    def show_degerlendirmeler_table(self):
        if self.islem.df_deg is None:
            if self.islem.df_ogrnotlar is None:
                messagebox.showerror(
                    "Hata", "Önce Öğrenci Notları tablosu oluşturun/yükleyin!"
                )
                return
            try:
                self.islem.create_deg_table_from_ogr()
                self.lbl_deg.config(text="Ders Çıktısı: (Yeni Oluşturuldu)")
                messagebox.showinfo(
                    "Bilgi",
                    "Ders Çıktısı tablosu, Öğrenci Notları sütunlarına göre oluşturuldu.",
                )
            except Exception as e:
                messagebox.showerror("Hata", str(e))
                return

        self._show_table_window(
            df_attr_name="df_deg",
            title="Ders Çıktısı Tablosu",
            add_button_text="Yeni Ders Çıktısı Ekle",
            save_callback=self.save_new_degerlendirme_row,
        )

    def show_prgciktisi_table(self):
        if self.islem.df_prgtablo is None:
            if (self.islem.df_deg is None) or (self.islem.df_deg.shape[0] == 0):
                messagebox.showerror(
                    "Hata",
                    "Önce Ders Çıktısı tablosu oluşturun ve en az 1 satır ekleyin!",
                )
                return
            try:
                self.islem.create_prg_table_from_deg()
                self.lbl_prg.config(
                    text="Program Çıktıları: (Yeni Oluşturuldu)")
                messagebox.showinfo(
                    "Bilgi",
                    "Program Çıktıları tablosu, Ders Çıktısı satırlarına göre oluşturuldu.",
                )
            except Exception as e:
                messagebox.showerror("Hata", str(e))
                return

        self._show_table_window(
            df_attr_name="df_prgtablo",
            title="Program Çıktıları Tablosu",
            add_button_text="Yeni Çıktı Ekle",
            save_callback=self.save_new_prg_row,
        )

    
    def _show_table_window(self, df_attr_name, title, add_button_text, save_callback):
        df = getattr(self.islem, df_attr_name)
        if df is None:
            return

        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("700x400")

        frame = tk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True)

        scroll = tk.Scrollbar(frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        tree = ttk.Treeview(
            frame, yscrollcommand=scroll.set, selectmode="browse")
        columns = list(df.columns)
        tree["columns"] = columns
        tree.column("#0", width=0, stretch=tk.NO)
        tree.heading("#0", text="", anchor=tk.CENTER)

        for col in columns:
            tree.column(col, anchor=tk.W, width=120)
            tree.heading(col, text=col, anchor=tk.CENTER)

        scroll.config(command=tree.yview)
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh():
            tree.delete(*tree.get_children())
            for _, row_data in df.iterrows():
                vals = tuple("" if pd.isna(v) else v for v in row_data)
                tree.insert("", tk.END, values=vals)

        refresh()

        btn_frame = tk.Frame(win)
        btn_frame.pack(fill=tk.X, pady=5)

        btn_add = tk.Button(
            btn_frame,
            text=add_button_text,
            command=lambda: self._open_add_window(
                df_attr_name, save_callback, refresh),
        )
        btn_add.pack(side=tk.LEFT, padx=10)

    def _open_add_window(self, df_attr_name, save_callback, refresh_func):
        add_win = tk.Toplevel(self)
        add_win.title("Yeni Satır Ekle")
        add_win.geometry("400x300")

        df = getattr(self.islem, df_attr_name)
        columns = list(df.columns)
        entry_vars = {}

        row_idx = 0
        for col in columns:
            lbl = tk.Label(add_win, text=col)
            lbl.grid(row=row_idx, column=0, padx=5, pady=5, sticky=tk.W)

            ent = tk.Entry(add_win)
            ent.grid(row=row_idx, column=1, padx=5, pady=5, sticky=tk.W)
            entry_vars[col] = ent
            row_idx += 1

        def on_save():
            new_data = {}
            for c in columns:
                new_data[c] = entry_vars[c].get()
            try:
                save_callback(df_attr_name, new_data)
            except Exception as e:
                messagebox.showerror("Hata", f"Satır eklenirken hata: {e}")
                return
            refresh_func()
            add_win.destroy()

        btn_save = tk.Button(add_win, text="Kaydet", command=on_save)
        btn_save.grid(row=row_idx, column=0, columnspan=2, pady=10)

    
    def save_new_student_row(self, df_attr_name, new_data):
        df = getattr(self.islem, df_attr_name)

        for i, col in enumerate(df.columns):
            if i == 0:
                pass
            else:
                val_str = new_data[col].strip()
                if val_str == "":
                    new_data[col] = float("nan")
                else:
                    val_float = float(val_str)
                    if val_float < 0 or val_float > 100:
                        raise ValueError(
                            f"'{col}' sütunundaki değer 0-100 aralığında olmalıdır.")
                    new_data[col] = val_float

        df.loc[len(df)] = new_data

        if self.ogr_file_path and os.path.exists(self.ogr_file_path):
            df.to_excel(self.ogr_file_path, index=False)
        else:
            df.to_excel("notlar_new.xlsx", index=False)

    def save_new_degerlendirme_row(self, df_attr_name, new_data):
        df = getattr(self.islem, df_attr_name)
        for i, col in enumerate(df.columns):
            if i == 0:
                pass
            else:
                val_str = new_data[col].strip()
                if val_str == "":
                    new_data[col] = float("nan")
                else:
                    val_float = float(val_str)
                    if val_float < 0 or val_float > 1:
                        raise ValueError(
                            f"'{col}' sütunundaki değer 0-1 aralığında olmalıdır.")
                    new_data[col] = val_float

        df.loc[len(df)] = new_data
        if self.deg_file_path and os.path.exists(self.deg_file_path):
            df.to_excel(self.deg_file_path, index=False)
        else:
            df.to_excel("degerlendirmeler_new.xlsx", startrow=1, index=False)

    def save_new_prg_row(self, df_attr_name, new_data):
        df = getattr(self.islem, df_attr_name)
        for i, col in enumerate(df.columns):
            if i == 0:
                pass
            else:
                val_str = new_data[col].strip()
                if val_str == "":
                    new_data[col] = float("nan")
                else:
                    val_float = float(val_str)
                    if val_float < 0 or val_float > 1:
                        raise ValueError(
                            f"'{col}' sütunundaki değer 0-1 aralığında olmalıdır.")
                    new_data[col] = val_float

        df.loc[len(df)] = new_data
        if self.prg_file_path and os.path.exists(self.prg_file_path):
            df.to_excel(self.prg_file_path, index=False)
        else:
            df.to_excel("progcikti_new.xlsx", startrow=1, index=False)

    
    def run_calculations(self):
        if (
            (self.islem.df_ogrnotlar is None)
            or (self.islem.df_deg is None)
            or (self.islem.df_prgtablo is None)
        ):
            messagebox.showwarning(
                "Uyarı",
                "Öğrenci Notları, Ders Çıktısı ve Program Çıktıları tablosu yok!"
            )
            return
        try:
            self.islem.tum_islemleri_yap()
            messagebox.showinfo(
                "Bilgi",
                "Hesaplamalar tamamlandı (Tablo3 hafızada, Tablo4 & Tablo5 de)."
            )
        except Exception as e:
            messagebox.showerror("Hata", f"Hesaplamalarda hata: {e}")

    def save_tables(self):
        """
        Kaydet butonuna basıldığında, Ders Adı metin kutusundan alınan değerle
        dersadi_tablo4.xlsx ve dersadi_tablo5.xlsx dosyaları oluşturulur.
        """
        try:
            ders_adi = self.txt_dersadi.get().strip()
            if not ders_adi:
                ders_adi = "dersadi"  # Boş kalırsa varsayılan bir isim

            tab4_file = f"{ders_adi}_tablo4.xlsx"
            tab5_file = f"{ders_adi}_tablo5.xlsx"

            self.islem.yaz_tablo4(tab4_file)
            self.islem.yaz_tablo5(tab5_file)

            messagebox.showinfo(
                "Bilgi",
                f"{tab4_file} ve {tab5_file} oluşturuldu."
            )
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt sırasında hata: {e}")


if __name__ == "__main__":
    app = DegerlendirmeApp()
    app.mainloop()
